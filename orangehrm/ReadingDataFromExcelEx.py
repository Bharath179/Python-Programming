import openpyxl

file="C:\\Users\\Lenovo\\Desktop\\Book2.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row    #count no of rows in excel sheet
cols=sheet.max_column  #count no of columns in excel sheet

#Reading all the rows and columns in excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='          ')
        print()

