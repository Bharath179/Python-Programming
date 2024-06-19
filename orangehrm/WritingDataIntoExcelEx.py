import openpyxl


file="C:\\Users\\Lenovo\\Desktop\\Book2.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=124
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=125
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="QA"

workbook.save(file)