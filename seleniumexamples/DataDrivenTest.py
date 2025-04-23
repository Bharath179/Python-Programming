import openpyxl
from openpyxl.styles import Border, Side

"""path="C:\\Users\\Lenovo\\Desktop\\datadriven.xlsx"

workbook=openpyxl.load_workbook(path)
#workbook.get_sheet_by_name("sheet1")
sheet=workbook.active
rows=sheet.max_row
column=sheet.max_column
print(rows)
print(column)

#Reading data from Excel
for row in range(1,rows+1):
    for column in range(1,column+1):
        print(sheet.cell(row=row,column=column).value,end=" ")
        
    print()
"""
#write data to Excel

"""path="C:\\Users\\Lenovo\\Desktop\\datadriven.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("Sheet2")
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value="welcome"
workbook.save(path)"""

#writing different sets of data

path = "C:\\Users\\Lenovo\\Desktop\\datadriven.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet2"]

data = [
    ["EmpName", "Designation", "Salary"],
    ["Alice", "Python Tester", "30k"],
    ["Bob", "Java Tester", "22k"],
    ["Charlie", "Java Developer", "35k"],
    ["Eve", "Python Developer", "28k"]
]

border_style = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

for r in range(1, 6):
    for c in range(1, 4):
        cell = sheet.cell(row=r, column=c)
        cell.value = data[r - 1][c - 1]
        cell.border = border_style
workbook.save(path)


